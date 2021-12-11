import datetime

import openpyxl

from Application.FlowReport.RowObjectData.RowObjectParser import RowObjectParser
from Processing.FlowReports.ExcelReporter.excelReporter import ExcelReport
from Application.FlowReport.TableNames.TableNames import TABLES_NAMES
from Utils.ConfigReader.ConfigReader import read_config
from Utils.DbConnector.dbConnector import DbConnector
from Utils.Logger.main_logger import get_logger

log = get_logger("FlowReport")


class FlowReport:
    def __init__(self):
        self.local_storage_path = read_config("local.storage").get("local_storage_path")


    def __read_sheet_data(self, report, report_name):
        if report_name == 'Актуальная выгрузка':
            ws = report['Актуальная выгрузка']
            table_name = 'items_history'
            col_max = 14
        elif report_name == 'Flow':
            ws = report['Flow']
            table_name = 'flow_sheet'
            col_max = 50
        elif report_name == 'Удаленные заказы':
            ws = report['Удаленные заказы']
            table_name = 'deleted_items_history'
            col_max = 15

        max_row = ws.max_row
        sheet_items_dict = {}

        for row in ws.iter_rows(min_row=1, max_row=max_row + 1, min_col=1, max_col=col_max):
            cur_row = []
            if row[0].value:
                for cell in row:
                    cur_row.append(cell.value)
                cur_row_object = RowObjectParser.parse(table_name=table_name,
                                                       row=cur_row,
                                                       items_dict=sheet_items_dict)
                log.debug(
                    f"""Read m_item_id from {report} sheet: {cur_row_object.m_item_id}""")
                cur_row_object.item_processing()
        return sheet_items_dict

    def __get_min_m_item_id_prev(self, items_dict):
        return min(items_dict)

    def __compare_attributes(self, values_new, values_prev):

        """ сравнение статусов изготовления изделий (шаг и останов)"""
        if (values_new.get('step_id') != values_prev.get("step_id") or
                values_new.get('m_item_stopped') != values_prev.get("m_item_stopped")):
            values_new.update({'m_item_step_status': 'Изменился статус изделия'})
        else:
            values_new.update({'m_item_step_status': ''})

        """ сравнение комментариев к изделию (comment и m_item_comment)"""

        if (values_new.get('comment') != values_prev.get("comment") or
                values_new.get('m_item_comment') != values_prev.get("m_item_comment")):
            values_new.update({'m_item_comment_status': 'Добавлен комментарий'})
            log.debug(f'result {values_new.get("m_item_comment_status")}')
        else:
            values_new.update({'m_item_comment_status': ''})
            log.debug(f'result {values_new.get("m_item_comment_status")}')

        """ сравнение размера изделия (m_item_size)"""
        if values_new.get('m_item_size') != values_prev.get("m_item_size"):
            values_new.update({'m_item_size_status': 'Изменился размер изделия'})
        else:
            values_new.update({'m_item_size_status': ''})

        """ сравнение цвета изделия (m_item_color)"""
        if values_new.get('m_item_color') != values_prev.get("m_item_color"):
            values_new.update({'m_item_color_status': 'Изменился цвет изделия'})
        else:
            values_new.update({'m_item_color_status': ''})

        """ сравнение подклада изделия (m_item_podklad)"""
        if values_new.get('m_item_podklad') != values_prev.get("m_item_podklad"):
            values_new.update({'m_item_podklad_status': 'Изменился подклад изделия'})
        else:
            values_new.update({'m_item_podklad_status': ''})

        """ сравнение модели изделия (m_item_model)"""
        if values_new.get('m_item_model') != values_prev.get("m_item_model"):
            values_new.update({'m_item_model_status': 'Изменилась модель изделия'})
        else:
            values_new.update({'m_item_model_status': ''})
        return values_new

    def __compare_items_dict(self, items_dict, prev_items_dict, deleted_items_dict):
        for m_item_id, values in items_dict.items():
            if m_item_id in prev_items_dict:
                """ статус изделия - учтено: проверить атрибуты на измненеия с предыдущей выгрузкой """
                values.update({'m_item_status': "учтено"})
                values = self.__compare_attributes(values_new=values, values_prev=prev_items_dict.get(m_item_id))
            else:
                """ статус изделия - новый, добавить на лист Актуальная выгрузка """
                values.update({'m_item_status': "новое изделие",
                               'm_item_step_status': '',
                               'm_item_comment_status': '',
                               'm_item_size_status': '',
                               'm_item_color_status': '',
                               'm_item_podklad_status': '',
                               'm_item_model_status': ''
                               })
            items_dict[m_item_id] = values

        for m_item_id, values in prev_items_dict.items():
            try:
                if int(m_item_id) < int(self.__get_min_m_item_id_prev(items_dict=items_dict)):
                    """ статус изделия - архивный, ничего не делаем """
                    pass
                else:
                    if m_item_id in items_dict:
                        log.debug("passing")
                    else:

                        """ статус изделия - удаленный, если отсутствует в удаленных изделиях, добавить на лист
                            удаленные заказы если есть, то ничего не делаем """
                        if m_item_id not in deleted_items_dict:
                            values.update({'m_item_delete_date': datetime.datetime.now().strftime("%d.%m.%Y")})
                            deleted_items_dict[m_item_id] = values

            except ValueError:
                continue

    def __get_flow_report(self, report_file_name):

        log.debug(f"report_path is {report_file_name}")
        report_data = openpyxl.load_workbook(filename=report_file_name, data_only=True)
        flow_sheet_dict = self.__read_sheet_data(report=report_data, report_name='Flow')

        flow_items_dict = {}
        for elem, val in flow_sheet_dict.items():
            log.debug(f"readed from flow {elem}")
            log.debug(f"readed from flow {val}")
            flow_items_dict[elem] = val

        report_data.close()

        report = openpyxl.load_workbook(filename=report_file_name)
        for table_name in TABLES_NAMES:
            connector = DbConnector(table_name=table_name, db_name="origin")
            items_list = connector.get_items_list()
            items_dict = {}
            log.info(f"Processing query result table {table_name}")
            for row in items_list:
                cur_row_object = RowObjectParser.parse(table_name=table_name, row=row, items_dict=items_dict)
                cur_row_object.item_processing()
            if table_name == 'items_history':
                prev_items_dict = self.__read_sheet_data(report=report, report_name='Актуальная выгрузка')
                deleted_items_dict = self.__read_sheet_data(report=report, report_name='Удаленные заказы')
                self.__compare_items_dict(items_dict=items_dict,
                                          prev_items_dict=prev_items_dict,
                                          deleted_items_dict=deleted_items_dict)
                excel_reporter = ExcelReport(items_dict=deleted_items_dict,
                                             report_name='deleted_items_history',
                                             report=report)
                excel_reporter.items_processing()
                del excel_reporter

                deleted_items_dict = self.__read_sheet_data(report=report, report_name='Удаленные заказы')
                excel_reporter = ExcelReport(items_dict=items_dict, report_name=table_name, report=report)
                excel_reporter.items_processing()
                max_key = max(items_dict)
                full_items_list = list(range(max_key + 1))
                full_items_list.pop(0)
                excel_reporter_flow = ExcelReport(items_dict=items_dict,
                                                  report_name='Flow',
                                                  report=report,
                                                  flow_sheet_dict=flow_sheet_dict,
                                                  full_items_list=full_items_list,
                                                  deleted_items_dict=deleted_items_dict)
                excel_reporter_flow.items_processing()

        report.save(filename=report_file_name)
        del items_list
        del excel_reporter_flow
        del deleted_items_dict
        del excel_reporter
        del report

        log.info("Recalculating values")
        # report_data = openpyxl.load_workbook(filename=report_file_name, data_only=True)
        # report.save(filename=report_file_name)
        return report_file_name

    def run_db_report(self, report_file_name):
        return self.__get_flow_report(report_file_name=report_file_name)
