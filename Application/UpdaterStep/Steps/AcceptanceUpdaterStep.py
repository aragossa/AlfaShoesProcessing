import datetime
import os

import openpyxl

from Application.UpdaterStep.Steps.Intervals import Intervals
from Application.UpdaterStep.UpdaterStep import UpdaterStep
from Utils.Logger.main_logger import get_logger

log = get_logger("AcceptanceUpdaterStep")


class AcceptanceUpdaterStep(UpdaterStep):

    def run_download_root_dir(self, processing_reports):
        log.info("Running Yandex Disk archiving")
        src_path = "/Учет Альфа/"
        files = self.ya.find_files_in_dir(src_path)
        filenames = [elem.name for elem in files]
        downloaded_files = []
        downloaded_files_clean = []
        for file in filenames:
            archive_report = False
            for report_name in processing_reports:
                if report_name in file:
                    archive_report = True
            if archive_report:
                log.debug(f"Processing file {file}")
                prepared_file_name = file.replace(".xlsx", "_архив.xlsx")
                src_file_path = f"{src_path}{file}"

                download_file_path = os.path.join(self.local_storage_path, prepared_file_name)
                self.ya.download_file(src_path=src_file_path,
                                      path_or_file=download_file_path)
                downloaded_files.append(download_file_path)
                downloaded_files_clean.append(file)

            else:
                log.debug(f"File {file} skipped")

        return downloaded_files, downloaded_files_clean

    def download_acceptance_archive(self):
        archive_dirs = self.ya.find_archive_dirs()
        acceptance_filepath = None
        acceptance_dirpath = None
        for curr_dir in archive_dirs:
            curr_file_name = self.ya.find_max_report(report_name='02_Приемка на склад', dir_name=curr_dir)
            if curr_file_name is not None:
                if acceptance_filepath is not None:
                    exist_file_change_date = f"{acceptance_filepath.split('_')[0]}_{acceptance_filepath.split('_')[1]}"
                    exist_this_datetime = datetime.datetime.strptime(exist_file_change_date, "%Y%m%d_%H%M")

                    found_file_change_date = f"{curr_file_name.split('_')[0]}_{curr_file_name.split('_')[1]}"
                    found_this_datetime = datetime.datetime.strptime(found_file_change_date, "%Y%m%d_%H%M")
                    if found_this_datetime > exist_this_datetime:
                        acceptance_filepath = curr_file_name
                        acceptance_dirpath = curr_dir
                else:
                    acceptance_filepath = curr_file_name
                    acceptance_dirpath = curr_dir
        src_file = f'{acceptance_dirpath}/{acceptance_filepath}'
        dst_file = os.path.join('TempFolder', acceptance_filepath)
        self.ya.download_file(src_path=src_file, path_or_file=dst_file)
        return dst_file

    def update_local_files(self, downloaded_files, new_files):
        for file in downloaded_files:
            if '02_Приемка на склад' in file:
                log.debug(f"Working with {file}")
                new_files.update({'02_Приемка на склад': file})
            elif '00_Справочник Альфа' in file:
                log.debug(f"Working with {file}")
                new_files.update({'00_Справочник Альфа': file})
            elif '00_Справочник Номенклатуры WB' in file:
                log.debug(f"Working with {file}")
                new_files.update({'00_Справочник Номенклатуры WB': file})
            elif '01_Flow' in file:
                log.debug(f"Working with {file}")
                new_files.update({'01_Flow': file})

    def download_acceptance_template(self):
        acceptance_template_dst_filepath = os.path.join(self.local_storage_path, self.acceptance_template_dst_filename)
        acceptance_template_src_filepath = 'Учет Альфа/Шаблоны учет Альфа/шаблон_02_Приемка на склад.xlsx'
        self.ya.download_file(acceptance_template_src_filepath, acceptance_template_dst_filepath)
        return acceptance_template_dst_filepath

    def prepare_acceptance_values(self, report_name, report_file_name, report):
        log.debug(f'reading {report_name}')
        read_values = Intervals.acceptance_intervals.get(report_name)
        log.debug(read_values)
        log.debug(f'opening {report_file_name}')
        data = openpyxl.load_workbook(filename=report_file_name, data_only=True, read_only=True)
        for sheet_name, params in read_values.items():
            log.debug(sheet_name)
            worksheet = data[sheet_name]
            if sheet_name == 'Flow':
                write_sheetname = 'Flow данные для приемки'
            else:
                write_sheetname = sheet_name
            report_sheet = report[write_sheetname]
            if params.get('cells') is not None:
                for cells in params.get('cells'):
                    log.debug(f'reading from {sheet_name}')
                    log.debug(cells)
                    value = self.read_cell_value(worksheet=worksheet,
                                                 col=cells.get('read').get('col'),
                                                 row=cells.get('read').get('row'))
                    log.debug(value)
                    log.debug(f'writing to {write_sheetname}')

                    self.write_cell_value(worksheet=report_sheet,
                                          col=cells.get('write').get('col'),
                                          row=cells.get('write').get('row'),
                                          value=value)
            for interval in params.get('intervals'):
                log.debug(f"current_interval")
                log.debug(f"{interval}")
                excel_data = self.read_interval(worksheet=worksheet,
                                                start_row=interval.get("read").get('start_row'),
                                                stop_row=interval.get("read").get('stop_row'),
                                                start_col=interval.get("read").get('start_col'),
                                                stop_col=interval.get("read").get('stop_col'))
                self.write_interval(worksheet=report_sheet,
                                    start_row=interval.get("write").get('start_row'),
                                    start_col=interval.get("write").get('start_col'),
                                    excel_data=excel_data)
        data.close()

    def read_cell_value(self, worksheet, col, row):
        log.debug('reading cell value')
        log.debug(f'{worksheet}, {col}, {row}')
        return worksheet.cell(row=row, column=col).value

    def write_cell_value(self, worksheet, col, row, value):
        log.debug('writing cell value')
        log.debug(f'{worksheet}, {col}, {row}, {value}')
        worksheet.cell(row=row, column=col).value = value

    def run_acceptance_updater(self, files, acceptance_report, acceptance_report_archive):
        log.debug(f'opening {acceptance_report}')
        report = openpyxl.load_workbook(filename=acceptance_report)
        log.debug(f'{files}')
        updated_file = files
        updated_file.update({'02_Приемка на склад': acceptance_report_archive})
        for report_name, report_file_name in files.items():
            self.prepare_acceptance_values(report_name=report_name, report_file_name=report_file_name, report=report)
        log.info(f'saving report {acceptance_report}')
        report.save(acceptance_report)
        report.close()

    def clean_root_dir(self, files):
        for file in files:
            if '02_Приемка на склад' in file:
                src_file = file.replace("TempFolder/", "")
                prepared_file_name = src_file.replace(".xlsx", "_архив.xlsx")
                destination_folder_name = prepared_file_name.split('_')[0]
                destination_path = f"/Учет Альфа/Архив учета Альфа/{destination_folder_name}/"
                destination_file_path = f"{destination_path}{prepared_file_name}"
                self.ya.mkdir(destination_path)
                self.ya.copy_file(src_path=f"/Учет Альфа/{src_file}",
                                  destination_file_path=destination_file_path,
                                  overwrite=True)

                self.ya.delete_file(src_path=f"/Учет Альфа/{src_file}")

    def upload_local_files(self, acceptance_report):
        clean_filename = acceptance_report.replace('TempFolder/', '').replace('TempFolder\\', '')
        self.ya.upload_file(src_path=acceptance_report,
                            destination_file_path=f"/Учет Альфа/{clean_filename}")
