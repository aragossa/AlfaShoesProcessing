import os

import openpyxl

from Application.UpdaterStep.UpdaterStep import UpdaterStep
from Utils.Logger.main_logger import get_logger
from Application.UpdaterStep.Steps.Intervals import Intervals

log = get_logger("RenewFlowUpdaterStep")


class RenewFlowUpdaterStep(UpdaterStep):

    def run_download_root_dir(self, processing_reports):
        log.info("Running Yandex Disk archiving")
        src_path = "/Учет Альфа/"
        files = self.ya.find_files_in_dir(src_path)
        filenames = [elem.name for elem in files]
        downloaded_files = []
        downloaded_files_clean = []
        for file in filenames:
            archive_report = False
            for report_name in processing_reports:
                if report_name in file:
                    archive_report = True
            if archive_report:
                log.debug(f"Processing file {file}")
                prepared_file_name = file.replace(".xlsx", "_архив.xlsx")
                src_file_path = f"{src_path}{file}"

                download_file_path = os.path.join(self.local_storage_path, prepared_file_name)
                self.ya.download_file(src_path=src_file_path,
                                      path_or_file=download_file_path)
                downloaded_files.append(download_file_path)
                downloaded_files_clean.append(file)

            else:
                log.debug(f"File {file} skipped")

        return downloaded_files, downloaded_files_clean

    def update_filename(self, file, new_files, report_name):
        log.debug(f'Updating report {report_name}')
        new_filename = f"{self.date_file_prefix}_{report_name}.xlsx"
        prepared_filename = os.path.join(self.local_storage_path, new_filename)
        os.rename(file, prepared_filename)
        new_files.update({report_name: prepared_filename})

    def update_local_files(self, downloaded_files, new_files):
        for file in downloaded_files:
            log.debug(f"Working with {file}")

            if '01_Flow' in file:
                self.update_filename(file=file,
                                     new_files=new_files,
                                     report_name='01_Flow')
            elif '02_Приемка на склад' in file:
                new_files.update({'02_Приемка на склад': file})

    def prepare_renew_flow_values(self, report_file_name, report):
        read_values = Intervals.renew_flow_intervals.get('02_Приемка на склад')
        log.debug(f'opening {report_file_name}')
        data = openpyxl.load_workbook(filename=report_file_name, data_only=True, read_only=True)
        for sheet_name, params in read_values.items():
            log.debug(sheet_name)
            worksheet = data[sheet_name]
            report_sheet = report['Сдано на склад']
            for interval in params.get('intervals'):
                excel_data = self.read_interval(worksheet=worksheet,
                                                start_row=interval.get('read').get('start_row'),
                                                stop_row=interval.get('read').get('stop_row'),
                                                start_col=interval.get('read').get('start_col'),
                                                stop_col=interval.get('read').get('stop_col'))
                log.debug(excel_data)
                self.write_interval(worksheet=report_sheet,
                                    start_row=interval.get('write').get('start_row'),
                                    start_col=interval.get('write').get('start_col'),
                                    excel_data=excel_data)
        data.close()

    def run_renew_flow_updater(self, files):
        log.debug(files)
        report = openpyxl.load_workbook(filename=files.get('01_Flow'))
        log.debug(f"opening {files.get('01_Flow')}")
        self.prepare_renew_flow_values(report_file_name=files.get('02_Приемка на склад'), report=report)
        log.info(f"saving report {files.get('01_Flow')}")
        report.save(files.get('01_Flow'))
        report.close()

    def clean_root_dir(self, files):
        for file in files:
            if '01_Flow' in file:
                src_file = file.replace("TempFolder/", "")
                self.ya.delete_file(src_path=f"/Учет Альфа/{src_file}")
            elif '02_Приемка на склад' in file:
                src_file = file.replace("TempFolder/", "")
                src_file_date = src_file.split('_')[0]
                dst_dir_name = f"/Учет Альфа/Архив учета Альфа/{src_file_date}"
                updated_file_name = src_file.replace('.xlsx', '_архив.xlsx')
                self.ya.mkdir(dst_dir_name)
                self.ya.move_file(src_path=f"/Учет Альфа/{src_file}",
                                  destination_file_path=f"{dst_dir_name}/{updated_file_name}")

    def upload_local_files(self, new_files):
        for report_type, file in new_files.items():
            if '01_Flow' in file:
                clean_filename = file.replace('TempFolder/', '').replace('TempFolder\\', '')
                self.ya.upload_file(src_path=file,
                                    destination_file_path=f"/Учет Альфа/{clean_filename}")

    def find_max_row(self, worksheet, start_row):
        max_row = None
        for row in worksheet.iter_rows(min_row=start_row):
            log.debug(row[0].value)
            log.debug(f'{row[0].value == None}')
            if row[0].value == None:
                max_row = row[0].row
                break
        if max_row is None:
            max_row = worksheet.max_row
        return max_row

    def write_interval(self, worksheet, start_row, start_col, excel_data):
        log.debug('writing intervals values')
        row = self.find_max_row(worksheet=worksheet, start_row=start_row)
        log.debug(f'found max row {row}')
        for rows in excel_data:
            log.debug(rows)
            col = start_col
            log.debug(f'writing to {worksheet}')
            log.debug(rows)
            for cell_value in rows:
                log.debug(f'writing to {worksheet} ({row}, {col}) value - {cell_value}')
                worksheet.cell(row=row, column=col).value = cell_value
                col += 1
            row += 1